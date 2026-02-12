"""
Data Analysis Service using Pandas and NumPy
Provides analytics and statistical analysis of student results
"""
import pandas as pd
import numpy as np
from typing import List, Dict, Optional
from sqlalchemy.orm import Session
from app.models import Student, Result, Subject, Semester
from app.schemas import SubjectStatistics, SemesterSummary
import logging
from openpyxl.styles import PatternFill, Font

logger = logging.getLogger(__name__)


class ResultAnalyzer:
    """
    Analyzes student results using pandas and numpy
    """

    def __init__(self, db: Session):
        self.db = db

    @staticmethod
    def _marks_to_percentage(total_marks: Optional[float]) -> Optional[float]:
        """Best-effort normalization of total marks to percentage.

        VTU PDFs vary by scheme; many totals are out of 100, some out of 150.
        We use a conservative heuristic to keep the system working until
        max-marks is explicitly modeled.
        """
        if total_marks is None or pd.isna(total_marks):
            return None

        try:
            m = float(total_marks)
        except Exception:
            return None

        if m < 0:
            return None

        # Heuristic: already looks like a percentage
        if m <= 100:
            return m
        # Common VTU totals
        if m <= 150:
            return (m / 150.0) * 100.0
        if m <= 200:
            return (m / 200.0) * 100.0

        # Fallback: treat as percentage-like to avoid breaking analytics
        return min(m, 100.0)

    @staticmethod
    def _percentage_to_grade_points(percent: Optional[float], result_status: Optional[str] = None) -> int:
        """Convert percentage to VTU CBCS grade points (10-point scale)."""
        if percent is None or pd.isna(percent):
            return 0

        # If explicitly failed/absent/etc, grade points should be 0
        if str(result_status or "").strip().upper() not in {"", "P", "PASS"}:
            return 0

        p = float(percent)
        if p >= 90:
            return 10  # O
        if p >= 80:
            return 9   # S
        if p >= 70:
            return 8   # A
        if p >= 60:
            return 7   # B
        if p >= 50:
            return 6   # C
        if p >= 45:
            return 5   # D
        if p >= 40:
            return 4   # E
        return 0       # F

    @staticmethod
    def _exam_month_to_number(exam_month: Optional[str]) -> Optional[int]:
        if exam_month is None:
            return None
        m = str(exam_month).strip().lower()
        mapping = {
            "january": 1,
            "jan": 1,
            "june": 6,
            "jun": 6,
            "july": 7,
            "jul": 7,
            "december": 12,
            "dec": 12,
        }
        return mapping.get(m)

    def _apply_carryover_fail_rule(self, df: pd.DataFrame) -> pd.DataFrame:
        """CBCS rule: after a student clears a previously failed course, award E grade (4 points).

        Best-effort ordering uses (exam_year, exam_month) when present.
        If term is unknown, any prior fail in DB triggers the override.
        """
        if df.empty:
            return df

        tmp = df.copy()

        # Normalize status
        tmp["_status"] = tmp["result_status"].astype(str).str.strip().str.upper()

        # Compute comparable term key
        tmp["_exam_month_num"] = tmp["exam_month"].apply(self._exam_month_to_number)
        tmp["_term_key"] = 0
        has_year = tmp["exam_year"].notna()
        has_month = tmp["_exam_month_num"].notna()
        tmp.loc[has_year & has_month, "_term_key"] = (
            tmp.loc[has_year & has_month, "exam_year"].astype(int) * 100
            + tmp.loc[has_year & has_month, "_exam_month_num"].astype(int)
        )

        key_cols = ["usn", "subject_id"] if "subject_id" in tmp.columns else ["usn", "subject_code"]

        # For each (usn, subject), compute earliest fail term, and whether any fail exists
        fail_rows = tmp[tmp["_status"].isin(["F", "FAIL"])].copy()
        any_fail = fail_rows.groupby(key_cols).size().rename("_any_fail")
        earliest_fail_term = (
            fail_rows.groupby(key_cols)["_term_key"].min().rename("_earliest_fail_term")
        )

        tmp = tmp.merge(any_fail, on=key_cols, how="left")
        tmp = tmp.merge(earliest_fail_term, on=key_cols, how="left")
        tmp["_any_fail"] = tmp["_any_fail"].fillna(0)
        tmp["_earliest_fail_term"] = tmp["_earliest_fail_term"].fillna(0)

        # Override rule applies only on passes
        is_pass = tmp["_status"].isin(["P", "PASS"])
        has_prior_fail_known_term = (tmp["_term_key"] > 0) & (tmp["_earliest_fail_term"] > 0) & (tmp["_earliest_fail_term"] < tmp["_term_key"])
        has_fail_unknown_term = (tmp["_term_key"] == 0) & (tmp["_any_fail"] > 0)
        should_override = is_pass & (has_prior_fail_known_term | has_fail_unknown_term)

        # Apply: E grade point = 4
        tmp.loc[should_override, "grade_points"] = 4

        # Cleanup helper columns
        tmp = tmp.drop(columns=[c for c in ["_status", "_exam_month_num", "_term_key", "_any_fail", "_earliest_fail_term"] if c in tmp.columns])
        return tmp

    def _compute_grade_points(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        tmp = df.copy()
        tmp["percent"] = tmp["total_marks"].apply(self._marks_to_percentage)
        tmp["grade_points"] = tmp.apply(
            lambda r: self._percentage_to_grade_points(r.get("percent"), r.get("result_status")), axis=1
        )
        # Apply CBCS carryover rule for previously failed courses
        tmp = self._apply_carryover_fail_rule(tmp)
        return tmp

    def get_semester_gpa(self, semester: int, batch: Optional[str] = None, branch: Optional[str] = None) -> pd.DataFrame:
        """Compute SGPA for each student in a given semester using CBCS credits."""
        df = self.get_results_dataframe(semester=semester, batch=batch, branch=branch)
        if df.empty:
            return pd.DataFrame()

        df = self._compute_grade_points(df)

        # Credits: treat 0 credits as non-credit (excluded). Missing credits default to 4 (common core).
        df["credits"] = df["credits"].fillna(4).astype(int)
        df = df[df["credits"] > 0]
        if df.empty:
            return pd.DataFrame()

        df["credit_points"] = df["credits"] * df["grade_points"]
        grouped = df.groupby(["usn", "student_name"]).agg(
            total_credits=("credits", "sum"),
            total_credit_points=("credit_points", "sum"),
        ).reset_index()

        grouped["sgpa"] = (grouped["total_credit_points"] / grouped["total_credits"]).round(2)
        return grouped

    def get_student_gpa_progression(self, usn: str) -> List[Dict]:
        """Return per-semester SGPA and running CGPA for a student."""
        df = self.get_results_dataframe(usn=usn)
        if df.empty:
            return []

        df = self._compute_grade_points(df)
        df["credits"] = df["credits"].fillna(4).astype(int)
        df = df[df["credits"] > 0]
        if df.empty:
            return []

        df["credit_points"] = df["credits"] * df["grade_points"]

        sem = df.groupby(["semester"]).agg(
            total_credits=("credits", "sum"),
            total_credit_points=("credit_points", "sum"),
        ).reset_index().sort_values("semester")

        sem["sgpa"] = (sem["total_credit_points"] / sem["total_credits"]).round(2)

        # Running CGPA = cumulative credit points / cumulative credits
        sem["cum_credits"] = sem["total_credits"].cumsum()
        sem["cum_credit_points"] = sem["total_credit_points"].cumsum()
        sem["cgpa"] = (sem["cum_credit_points"] / sem["cum_credits"]).round(2)

        return [
            {
                "semester": int(r.semester),
                "total_credits": int(r.total_credits),
                "sgpa": float(r.sgpa),
                "cgpa": float(r.cgpa),
            }
            for r in sem.itertuples(index=False)
        ]

    def get_results_dataframe(
        self,
        semester: Optional[int] = None,
        usn: Optional[str] = None,
        batch: Optional[str] = None,
        branch: Optional[str] = None,
        exam_year: Optional[int] = None,
        exam_month: Optional[str] = None,
    ) -> pd.DataFrame:
        """
        Get results as a pandas DataFrame with filters
        
        Args:
            semester: Filter by semester number
            usn: Filter by student USN
            
        Returns:
            DataFrame with result data
        """
        try:
            query = self.db.query(
                Student.usn,
                Student.student_name,
                Student.batch,
                Student.branch,
                Semester.semester_number,
                Semester.exam_month,
                Semester.exam_year,
                Subject.id,
                Subject.subject_code,
                Subject.subject_name,
                Subject.credits,
                Result.internal_marks,
                Result.external_marks,
                Result.total_marks,
                Result.result_status,
                Result.announced_date
            ).join(Result.student).join(Result.semester).join(Result.subject)

            if semester:
                query = query.filter(Semester.semester_number == semester)
            if usn:
                query = query.filter(Student.usn == usn)
            if batch:
                query = query.filter(Student.batch == batch)
            if branch:
                query = query.filter(Student.branch == branch)
            if exam_year is not None:
                query = query.filter(Semester.exam_year == exam_year)
            if exam_month:
                query = query.filter(Semester.exam_month.ilike(exam_month))

            results = query.all()

            # Convert to DataFrame
            df = pd.DataFrame(results, columns=[
                'usn', 'student_name', 'batch', 'branch', 'semester', 'exam_month', 'exam_year',
                'subject_id', 'subject_code', 'subject_name', 'credits', 'internal_marks', 'external_marks',
                'total_marks', 'result_status', 'announced_date'
            ])

            return df

        except Exception as e:
            logger.error(f"Error creating DataFrame: {str(e)}")
            return pd.DataFrame()

    def get_subject_statistics(
        self,
        semester: int,
        batch: Optional[str] = None,
        branch: Optional[str] = None,
        exam_year: Optional[int] = None,
        exam_month: Optional[str] = None,
    ) -> List[SubjectStatistics]:
        """
        Calculate statistics for each subject in a semester
        
        Args:
            semester: Semester number
            
        Returns:
            List of SubjectStatistics
        """
        df = self.get_results_dataframe(semester=semester, batch=batch, branch=branch, exam_year=exam_year, exam_month=exam_month)
        
        if df.empty:
            return []

        # Group by subject
        stats = []
        for subject_code in df['subject_code'].unique():
            subject_df = df[df['subject_code'] == subject_code]
            
            stat = SubjectStatistics(
                subject_code=subject_code,
                subject_name=subject_df['subject_name'].iloc[0],
                total_students=len(subject_df),
                avg_internal=float(subject_df['internal_marks'].mean()) if not subject_df['internal_marks'].isna().all() else None,
                avg_external=float(subject_df['external_marks'].mean()) if not subject_df['external_marks'].isna().all() else None,
                avg_total=float(subject_df['total_marks'].mean()) if not subject_df['total_marks'].isna().all() else None,
                max_marks=int(subject_df['total_marks'].max()) if not subject_df['total_marks'].isna().all() else None,
                min_marks=int(subject_df['total_marks'].min()) if not subject_df['total_marks'].isna().all() else None,
                pass_count=int((subject_df['result_status'] == 'P').sum()),
                fail_count=int((subject_df['result_status'] == 'F').sum()),
                pass_percentage=float((subject_df['result_status'] == 'P').sum() / len(subject_df) * 100) if len(subject_df) > 0 else None
            )
            stats.append(stat)

        return stats

    def get_student_summary(self, usn: str) -> List[SemesterSummary]:
        """
        Get semester-wise summary for a student
        
        Args:
            usn: Student USN
            
        Returns:
            List of SemesterSummary
        """
        df = self.get_results_dataframe(usn=usn)
        
        if df.empty:
            return []

        summaries = []
        for semester in df['semester'].unique():
            sem_df = df[df['semester'] == semester]
            
            summary = SemesterSummary(
                usn=usn,
                student_name=sem_df['student_name'].iloc[0],
                semester_number=int(semester),
                exam_month=sem_df['exam_month'].iloc[0] if not sem_df['exam_month'].isna().all() else None,
                exam_year=int(sem_df['exam_year'].iloc[0]) if not sem_df['exam_year'].isna().all() else None,
                total_subjects=len(sem_df),
                subjects_passed=int((sem_df['result_status'] == 'P').sum()),
                subjects_failed=int((sem_df['result_status'] == 'F').sum()),
                average_marks=float(sem_df['total_marks'].mean()) if not sem_df['total_marks'].isna().all() else None,
                highest_marks=int(sem_df['total_marks'].max()) if not sem_df['total_marks'].isna().all() else None,
                lowest_marks=int(sem_df['total_marks'].min()) if not sem_df['total_marks'].isna().all() else None
            )
            summaries.append(summary)

        return sorted(summaries, key=lambda x: x.semester_number)

    def get_top_performers(
        self,
        semester: int,
        subject_code: Optional[str] = None,
        limit: int = 10,
        rank_by: str = "marks",
        batch: Optional[str] = None,
        branch: Optional[str] = None,
        exam_year: Optional[int] = None,
        exam_month: Optional[str] = None,
    ) -> pd.DataFrame:
        """
        Get top performing students
        
        Args:
            semester: Semester number
            subject_code: Optional subject code filter
            limit: Number of top performers to return
            
        Returns:
            DataFrame with top performers
        """
        df = self.get_results_dataframe(semester=semester, batch=batch, branch=branch, exam_year=exam_year, exam_month=exam_month)
        
        if df.empty:
            return pd.DataFrame()

        if subject_code:
            # For specific subject, return top marks in that subject
            df = df[df['subject_code'] == subject_code]
            top_df = df.nlargest(limit, 'total_marks')
            return top_df[['usn', 'student_name', 'subject_code', 'subject_name', 'total_marks', 'result_status']]
        else:
            rank_by_norm = (rank_by or "marks").strip().lower()

            if rank_by_norm == "sgpa":
                gpa_df = self.get_semester_gpa(semester, batch=batch, branch=branch)
                if gpa_df.empty:
                    return pd.DataFrame()
                top_df = gpa_df.nlargest(limit, "sgpa")
                top_df["subject_code"] = "ALL"
                top_df["subject_name"] = f"Semester {semester} SGPA"
                top_df["total_marks"] = None
                top_df["result_status"] = None
                return top_df[["usn", "student_name", "subject_code", "subject_name", "total_marks", "result_status", "sgpa", "total_credits"]]

            # Default: aggregate total marks per student (existing behavior)
            student_totals = df.groupby(['usn', 'student_name']).agg({
                'total_marks': 'sum',
                'result_status': lambda x: 'P' if (x == 'P').all() else 'F'
            }).reset_index()

            top_df = student_totals.nlargest(limit, 'total_marks')
            top_df['subject_code'] = 'ALL'
            top_df['subject_name'] = f'Semester {semester} Aggregate'
            return top_df[['usn', 'student_name', 'subject_code', 'subject_name', 'total_marks', 'result_status']]

    def get_failure_analysis(
        self,
        semester: int,
        batch: Optional[str] = None,
        branch: Optional[str] = None,
        exam_year: Optional[int] = None,
        exam_month: Optional[str] = None,
    ) -> Dict:
        """
        Analyze failures in a semester
        
        Args:
            semester: Semester number
            
        Returns:
            Dictionary with failure statistics
        """
        df = self.get_results_dataframe(semester=semester, batch=batch, branch=branch, exam_year=exam_year, exam_month=exam_month)
        
        if df.empty:
            return {}

        failed_df = df[df['result_status'] == 'F']
        
        analysis = {
            'total_results': len(df),
            'total_failures': len(failed_df),
            'failure_rate': float(len(failed_df) / len(df) * 100) if len(df) > 0 else 0,
            'subject_wise_failures': failed_df.groupby('subject_code')['subject_name'].count().to_dict(),
            'students_with_failures': failed_df['usn'].unique().tolist()
        }

        return analysis

    def get_overall_statistics(self, batch: Optional[str] = None, branch: Optional[str] = None) -> Dict:
        """
        Get overall statistics across all results
        
        Returns:
            Dictionary with overall statistics
        """
        df = self.get_results_dataframe(batch=batch, branch=branch)
        
        if df.empty:
            return {
                'pass_rate': 0.0,
                'total_students': 0,
                'total_records': 0,
                'average_cgpa': 0.0
            }

        # Calculate pass rate
        total_results = len(df)
        passed_results = len(df[df['result_status'] == 'P'])
        pass_rate = (passed_results / total_results * 100) if total_results > 0 else 0.0

        # Count unique students
        total_students = df['usn'].nunique()

        # Calculate CGPA using VTU grading system (10-point scale)
        # VTU Grade Points based on total marks (out of 100)
        def marks_to_grade_points(marks):
            """Convert marks to VTU grade points"""
            if pd.isna(marks):
                return 0
            # Assuming marks are out of 100 (normalized)
            # If total_marks is out of 150 or other max, we need to normalize
            # For now, assuming it's percentage or out of 100
            if marks >= 90: return 10  # O - Outstanding
            elif marks >= 80: return 9  # S - Excellent
            elif marks >= 70: return 8  # A - Very Good
            elif marks >= 60: return 7  # B - Good
            elif marks >= 55: return 6  # C - Fair
            elif marks >= 50: return 5  # D - Satisfactory
            elif marks >= 40: return 4  # P/E - Pass
            else: return 0              # F - Fail

        # Assuming equal credits for all subjects (typical is 3-4 credits per subject)
        # Since we don't have credits in DB, we'll calculate simple average of grade points
        df['grade_points'] = df['total_marks'].apply(marks_to_grade_points)
        
        # Calculate CGPA per student (average grade points across all subjects)
        # Filter out only passed subjects for CGPA or include all (VTU includes all)
        student_cgpa = df.groupby('usn')['grade_points'].mean()
        average_cgpa = student_cgpa.mean() if len(student_cgpa) > 0 else 0.0

        return {
            'pass_rate': round(pass_rate, 2),
            'total_students': int(total_students),
            'total_records': int(total_results),
            'average_cgpa': round(average_cgpa, 2)
        }

    def compare_semesters(self, usn: str) -> pd.DataFrame:
        """
        Compare performance across semesters for a student
        
        Args:
            usn: Student USN
            
        Returns:
            DataFrame with semester-wise comparison
        """
        df = self.get_results_dataframe(usn=usn)
        
        if df.empty:
            return pd.DataFrame()

        # Group by semester and calculate metrics
        comparison = df.groupby('semester').agg({
            'total_marks': ['mean', 'max', 'min', 'std'],
            'subject_code': 'count',
            'result_status': lambda x: (x == 'P').sum()
        }).round(2)

        comparison.columns = ['avg_marks', 'max_marks', 'min_marks', 'std_dev', 'total_subjects', 'subjects_passed']
        
        return comparison

    def export_to_excel(
        self,
        output_path: str,
        semester: Optional[int] = None,
        batch: Optional[str] = None,
        branch: Optional[str] = None,
    ):
        """
        Export results to Excel file with multiple sheets
        
        Args:
            output_path: Path to save Excel file
            semester: Optional semester filter
        """
        try:
            df = self.get_results_dataframe(semester=semester, batch=batch, branch=branch)
            
            if df.empty:
                logger.warning("No data to export")
                return

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Sheet 1: Student-wise View (Pivot Format) - Easy to read
                student_pivot = self._create_student_pivot(df)
                student_pivot.to_excel(writer, sheet_name='Student View', index=False)
                
                # Sheet 2: Raw Data with unique IDs - For detailed tracking
                df_with_id = df.copy()
                # Add unique result_id as first column
                df_with_id.insert(0, 'result_id', df_with_id.index + 1)
                df_with_id.to_excel(writer, sheet_name='Raw Data', index=False)
                
                # Sheet 3: Subject Statistics (if semester specified)
                if semester:
                    stats = self.get_subject_statistics(semester, batch=batch, branch=branch)
                    stats_df = pd.DataFrame([s.model_dump() for s in stats])
                    stats_df.to_excel(writer, sheet_name='Subject Statistics', index=False)
                
                # Auto-size columns and apply conditional formatting for all sheets
                red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                red_font = Font(color="DC2626", bold=True)
                
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Auto-size columns
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Highlight failed students in red
                    if sheet_name == 'Student View':
                        # Find Subjects_Failed column
                        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=False))
                        failed_col_idx = None
                        for idx, cell in enumerate(header_row, 1):
                            if cell.value == 'Subjects_Failed':
                                failed_col_idx = idx
                                break
                        
                        # Highlight rows where Subjects_Failed > 0
                        if failed_col_idx:
                            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                                failed_count = worksheet.cell(row=row_idx, column=failed_col_idx).value
                                if failed_count and int(failed_count) > 0:
                                    for cell in row:
                                        cell.fill = red_fill
                                        cell.font = red_font
                    
                    elif sheet_name == 'Raw Data':
                        # Find result_status column
                        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=False))
                        status_col_idx = None
                        for idx, cell in enumerate(header_row, 1):
                            if cell.value == 'result_status':
                                status_col_idx = idx
                                break
                        
                        # Highlight rows where result_status = 'F'
                        if status_col_idx:
                            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                                status = worksheet.cell(row=row_idx, column=status_col_idx).value
                                if status == 'F':
                                    for cell in row:
                                        cell.fill = red_fill
                                        cell.font = red_font

            logger.info(f"Exported data to {output_path}")

        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise

    def _create_student_pivot(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Create pivot format: one row per student with subjects as columns
        
        Args:
            df: Results dataframe
            
        Returns:
            Pivot dataframe with student-wise view
        """
        if df.empty:
            return pd.DataFrame()
        
        pivot_data = []
        
        # Group by student and semester
        for (usn, student_name, batch, branch, semester), group in df.groupby(['usn', 'student_name', 'batch', 'branch', 'semester']):
            row = {
                'USN': usn,
                'Student Name': student_name,
                'Batch': batch,
                'Branch': branch,
                'Semester': semester,
            }
            
            # Add exam details if available
            if 'exam_month' in group.columns and not group['exam_month'].isna().all():
                row['Exam Month'] = group['exam_month'].iloc[0]
            if 'exam_year' in group.columns and not group['exam_year'].isna().all():
                row['Exam Year'] = int(group['exam_year'].iloc[0])
            
            # Add each subject's marks
            for idx, result in group.iterrows():
                subject_code = result['subject_code']
                row[f'{subject_code}_Internal'] = result['internal_marks']
                row[f'{subject_code}_External'] = result['external_marks']
                row[f'{subject_code}_Total'] = result['total_marks']
                row[f'{subject_code}_Result'] = result['result_status']
            
            # Calculate summary statistics
            row['Total_Marks'] = group['total_marks'].sum()
            row['Average_Marks'] = round(group['total_marks'].mean(), 2)
            row['Subjects_Passed'] = (group['result_status'] == 'P').sum()
            row['Subjects_Failed'] = (group['result_status'] == 'F').sum()
            row['Total_Subjects'] = len(group)
            
            # Calculate SGPA (Semester Grade Point Average)
            def marks_to_gp(marks):
                if pd.isna(marks): return 0
                if marks >= 90: return 10
                elif marks >= 80: return 9
                elif marks >= 70: return 8
                elif marks >= 60: return 7
                elif marks >= 55: return 6
                elif marks >= 50: return 5
                elif marks >= 40: return 4
                else: return 0
            
            grade_points = group['total_marks'].apply(marks_to_gp)
            row['SGPA'] = round(grade_points.mean(), 2)
            
            pivot_data.append(row)
        
        return pd.DataFrame(pivot_data)

    def export_to_csv(
        self,
        output_path: str,
        semester: Optional[int] = None,
        batch: Optional[str] = None,
        branch: Optional[str] = None,
    ):
        """
        Export results to CSV file in pivot format (student-wise view)
        
        Args:
            output_path: Path to save CSV file
            semester: Optional semester filter
        """
        try:
            df = self.get_results_dataframe(semester=semester, batch=batch, branch=branch)
            
            if df.empty:
                logger.warning("No data to export")
                return

            # Create student-wise pivot view for CSV
            student_pivot = self._create_student_pivot(df)
            student_pivot.to_csv(output_path, index=False, encoding='utf-8')
            logger.info(f"Exported data to {output_path}")

        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            raise
 
 
